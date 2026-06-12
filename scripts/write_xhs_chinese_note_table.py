import json
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill


HEADERS = [
    "品类信息",
    "笔记链接地址",
    "笔记作者名称",
    "内容概述",
    "标签信息",
    "粉丝量",
    "点赞量",
    "合作风险",
]


def main() -> None:
    if len(sys.argv) != 3:
        raise SystemExit("usage: write_xhs_chinese_note_table.py rows.json output.xlsx")

    rows_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2])
    rows = json.loads(rows_path.read_text(encoding="utf-8"))

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    usable_rows = [
        item for item in rows
        if not str(item.get("合作风险", "")).startswith("高风险")
    ]
    risk_rows = [
        item for item in rows
        if str(item.get("合作风险", "")).startswith("高风险")
    ]
    write_sheet(wb, "中文素材", usable_rows)
    if risk_rows:
        write_sheet(wb, "高风险复核", risk_rows)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def write_sheet(wb, title: str, rows: list[dict]) -> None:
    ws = wb.create_sheet(title[:31])
    header_fill = PatternFill("solid", fgColor="D9EAD3")
    risk_fill = PatternFill("solid", fgColor="F4CCCC")
    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(1, col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for row_idx, item in enumerate(rows, start=2):
        is_high_risk = str(item.get("合作风险", "")).startswith("高风险")
        for col_idx, header in enumerate(HEADERS, start=1):
            cell = ws.cell(row_idx, col_idx)
            cell.value = item.get(header, "")
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if is_high_risk:
                cell.fill = risk_fill

    widths = {
        "A": 22,
        "B": 60,
        "C": 20,
        "D": 42,
        "E": 46,
        "F": 12,
        "G": 12,
        "H": 28,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


if __name__ == "__main__":
    main()
